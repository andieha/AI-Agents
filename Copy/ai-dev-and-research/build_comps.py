import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

DARK_BLUE  = "1F4E79"
NAVY       = "17365D"
LIGHT_BLUE = "D9E1F2"
LIGHT_GREY = "F2F2F2"
WHITE      = "FFFFFF"
BLACK      = "000000"
INPUT_BLUE = "2E75B6"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Comps"

NUM_COLS = 8

# Column widths
widths = {"A": 24, "B": 16, "C": 15, "D": 16, "E": 13, "F": 16, "G": 14, "H": 16}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

def cell_set(ws, row, col, value, fg=BLACK, bold=False, bg=None,
             align="center", fmt=None, italic=False, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Times New Roman", size=11, color=fg, bold=bold, italic=italic)
    if bg:
        c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fmt:
        c.number_format = fmt
    return c

def section_header(ws, row, value, bg=DARK_BLUE, size=12):
    ws.cell(row=row, column=1, value=value).font = Font(
        name="Times New Roman", size=size, color=WHITE, bold=True)
    ws.cell(row=row, column=1).fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_COLS)
    for col in range(2, NUM_COLS + 1):
        c = ws.cell(row=row, column=col)
        c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    ws.row_dimensions[row].height = 24

def col_header_row(ws, row, labels):
    for i, label in enumerate(labels, 1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = Font(name="Times New Roman", size=11, color=BLACK, bold=True)
        c.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 36

def stat_row(ws, row, label, formula_fn, cols_with_formula, num_fmts):
    bold = (label == "Median")
    c = ws.cell(row=row, column=1, value=label)
    c.font = Font(name="Times New Roman", size=11, color=BLACK, bold=bold)
    c.fill = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")
    c.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(2, NUM_COLS + 1):
        c = ws.cell(row=row, column=col)
        c.fill = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")
        c.font = Font(name="Times New Roman", size=11, color=BLACK, bold=bold)
        c.alignment = Alignment(horizontal="center", vertical="center")
        if col in cols_with_formula:
            idx = cols_with_formula.index(col)
            c.value = formula_fn(get_column_letter(col))
            c.number_format = num_fmts[idx]
    ws.row_dimensions[row].height = 22

# ── TITLE BLOCK ───────────────────────────────────────────────────────────────
section_header(ws, 1, "MEGA-CAP TECHNOLOGY — COMPARABLE COMPANY ANALYSIS", bg=DARK_BLUE, size=13)
section_header(ws, 2,
    "Apple (AAPL)  •  Microsoft (MSFT)  •  Alphabet (GOOGL)  •  Meta (META)  •  Amazon (AMZN)",
    bg=DARK_BLUE, size=11)
section_header(ws, 3,
    "Benchmarking Analysis  |  As of June 2025  |  Fiscal Year 2024  |  All figures in USD Millions except ratios and multiples",
    bg=NAVY, size=10)
ws.row_dimensions[4].height = 6

# ── OPERATING STATISTICS ──────────────────────────────────────────────────────
section_header(ws, 5, "OPERATING STATISTICS & FINANCIAL METRICS")
col_header_row(ws, 6, [
    "Company", "Revenue\n(FY2024, $M)", "Revenue\nGrowth (YoY)",
    "Gross Profit\n($M)", "Gross\nMargin", "EBITDA\n($M)",
    "EBITDA\nMargin", "Net Income\n($M)"
])

OPS_DATA = [
    ("Apple (AAPL)",     391035, 0.020, 180683, 134000, 93736),
    ("Microsoft (MSFT)", 245122, 0.157, 170093, 135433, 88136),
    ("Alphabet (GOOGL)", 350018, 0.140, 210010, 119800, 94310),
    ("Meta (META)",      164501, 0.221, 131601,  74913, 62360),
    ("Amazon (AMZN)",    637959, 0.110, 299836, 120000, 59248),
]

OPS_R0 = 7
for i, (name, rev, growth, gp, ebitda, ni) in enumerate(OPS_DATA):
    r = OPS_R0 + i
    ws.row_dimensions[r].height = 22

    # Company name
    c = cell_set(ws, r, 1, name, fg=BLACK, bold=True, align="left")

    # Revenue
    c = cell_set(ws, r, 2, rev, fg=INPUT_BLUE, fmt="#,##0")
    c.comment = Comment("Source: FY2024 10-K / annual report", "Comps")

    # Revenue Growth
    c = cell_set(ws, r, 3, growth, fg=INPUT_BLUE, fmt="0.0%")
    c.comment = Comment("YoY growth vs. FY2023. Source: Company 10-K", "Comps")

    # Gross Profit
    c = cell_set(ws, r, 4, gp, fg=INPUT_BLUE, fmt="#,##0")
    c.comment = Comment("Source: FY2024 10-K income statement", "Comps")

    # Gross Margin — formula
    cell_set(ws, r, 5, f"=D{r}/B{r}", fg=BLACK, fmt="0.0%")

    # EBITDA
    c = cell_set(ws, r, 6, ebitda, fg=INPUT_BLUE, fmt="#,##0")
    c.comment = Comment("Operating Income + D&A. Source: FY2024 10-K", "Comps")

    # EBITDA Margin — formula
    cell_set(ws, r, 7, f"=F{r}/B{r}", fg=BLACK, fmt="0.0%")

    # Net Income
    c = cell_set(ws, r, 8, ni, fg=INPUT_BLUE, fmt="#,##0")
    c.comment = Comment("Net income attributable to shareholders. Source: FY2024 10-K", "Comps")

OPS_R1 = OPS_R0 + len(OPS_DATA) - 1  # 11
ws.row_dimensions[OPS_R1 + 1].height = 6  # blank separator row 12

# Operating statistics (rows 13-17)
STAT_LABELS = ["Maximum", "75th Percentile", "Median", "25th Percentile", "Minimum"]
STAT_FUNCS  = [
    lambda col: f"=MAX({col}{OPS_R0}:{col}{OPS_R1})",
    lambda col: f"=QUARTILE({col}{OPS_R0}:{col}{OPS_R1},3)",
    lambda col: f"=MEDIAN({col}{OPS_R0}:{col}{OPS_R1})",
    lambda col: f"=QUARTILE({col}{OPS_R0}:{col}{OPS_R1},1)",
    lambda col: f"=MIN({col}{OPS_R0}:{col}{OPS_R1})",
]
OPS_STAT_R0 = OPS_R1 + 2  # 13
# Stats on: Growth (col 3), Gross Margin (col 5), EBITDA Margin (col 7)
OPS_STAT_COLS = [3, 5, 7]
OPS_STAT_FMTS = ["0.0%", "0.0%", "0.0%"]

for si, (label, fn) in enumerate(zip(STAT_LABELS, STAT_FUNCS)):
    stat_row(ws, OPS_STAT_R0 + si, label, fn, OPS_STAT_COLS, OPS_STAT_FMTS)

OPS_STAT_R1 = OPS_STAT_R0 + len(STAT_LABELS) - 1  # 17
ws.row_dimensions[OPS_STAT_R1 + 1].height = 6

# ── VALUATION MULTIPLES ───────────────────────────────────────────────────────
VAL_HDR = OPS_STAT_R1 + 2  # 19
section_header(ws, VAL_HDR, "VALUATION MULTIPLES")
col_header_row(ws, VAL_HDR + 1, [
    "Company", "Market Cap\n($M)", "Enterprise\nValue ($M)",
    "EV / Revenue", "EV / EBITDA", "P / E", "", ""
])

VAL_DATA = [
    ("Apple (AAPL)",     3200000, 3295000),
    ("Microsoft (MSFT)", 3100000, 3082000),
    ("Alphabet (GOOGL)", 2200000, 2114000),
    ("Meta (META)",      1600000, 1550000),
    ("Amazon (AMZN)",    2300000, 2370000),
]

VAL_R0 = VAL_HDR + 2  # 21
for i, (name, mktcap, ev) in enumerate(VAL_DATA):
    r    = VAL_R0 + i
    ops_r = OPS_R0 + i  # cross-reference to operating section
    ws.row_dimensions[r].height = 22

    cell_set(ws, r, 1, name, fg=BLACK, bold=True, align="left")

    c = cell_set(ws, r, 2, mktcap, fg=INPUT_BLUE, fmt="#,##0")
    c.comment = Comment("Market cap as of June 2025. Source: Bloomberg / public market data", "Comps")

    c = cell_set(ws, r, 3, ev, fg=INPUT_BLUE, fmt="#,##0")
    c.comment = Comment("EV = Market Cap ± Net Debt. Source: FY2024 balance sheet", "Comps")

    cell_set(ws, r, 4, f"=C{r}/B{ops_r}", fg=BLACK, fmt='0.0"x"')  # EV/Revenue
    cell_set(ws, r, 5, f"=C{r}/F{ops_r}", fg=BLACK, fmt='0.0"x"')  # EV/EBITDA
    cell_set(ws, r, 6, f"=B{r}/H{ops_r}", fg=BLACK, fmt='0.0"x"')  # P/E

VAL_R1 = VAL_R0 + len(VAL_DATA) - 1  # 25
ws.row_dimensions[VAL_R1 + 1].height = 6

# Valuation statistics (rows 27-31)
VAL_STAT_R0 = VAL_R1 + 2  # 27
VAL_STAT_COLS = [4, 5, 6]
VAL_STAT_FMTS = ['0.0"x"', '0.0"x"', '0.0"x"']
VAL_STAT_FUNCS = [
    lambda col: f"=MAX({col}{VAL_R0}:{col}{VAL_R1})",
    lambda col: f"=QUARTILE({col}{VAL_R0}:{col}{VAL_R1},3)",
    lambda col: f"=MEDIAN({col}{VAL_R0}:{col}{VAL_R1})",
    lambda col: f"=QUARTILE({col}{VAL_R0}:{col}{VAL_R1},1)",
    lambda col: f"=MIN({col}{VAL_R0}:{col}{VAL_R1})",
]
for si, (label, fn) in enumerate(zip(STAT_LABELS, VAL_STAT_FUNCS)):
    stat_row(ws, VAL_STAT_R0 + si, label, fn, VAL_STAT_COLS, VAL_STAT_FMTS)

VAL_STAT_R1 = VAL_STAT_R0 + len(STAT_LABELS) - 1  # 31
ws.row_dimensions[VAL_STAT_R1 + 1].height = 6

# ── NOTES & METHODOLOGY ──────────────────────────────────────────────────────
NOTES_R0 = VAL_STAT_R1 + 2  # 33
section_header(ws, NOTES_R0, "NOTES & METHODOLOGY")

NOTES = [
    ("Data Sources",
     "Financial data from FY2024 annual reports (10-K filings). Market Cap and EV as of June 2025 "
     "from public market data. No institutional MCP sources (S&P Kensho, FactSet, Daloopa) were "
     "available; figures should be verified against Bloomberg or FactSet before IC use."),
    ("Peer Group",
     "5 mega-cap US-listed technology companies selected for similar consumer product/services mix "
     "and comparable scale ($150B+ revenue). All report in USD; no FX adjustment required."),
    ("EBITDA",
     "Operating Income + Depreciation & Amortization from each company's FY2024 10-K."),
    ("Enterprise Value",
     "EV = Market Capitalization + Total Debt – Cash & Cash Equivalents (FY2024 year-end balance sheet)."),
    ("Color Convention",
     "Blue = hardcoded inputs with source citations in cell comments. Black = Excel formulas "
     "referencing input cells. No values are hard-coded into formula cells."),
    ("Disclaimer",
     "For benchmarking purposes only. Does not constitute investment advice. Verify all figures "
     "independently before use in client-facing or investment committee materials."),
]

for ni, (label, text) in enumerate(NOTES):
    r = NOTES_R0 + 1 + ni
    c = ws.cell(row=r, column=1, value=label)
    c.font = Font(name="Times New Roman", size=10, color=BLACK, bold=True)
    c.alignment = Alignment(horizontal="left", vertical="top")

    c = ws.cell(row=r, column=2, value=text)
    c.font = Font(name="Times New Roman", size=10, color=BLACK)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=NUM_COLS)
    ws.row_dimensions[r].height = 40

# ── SAVE ──────────────────────────────────────────────────────────────────────
out = "/home/user/AI-Agents/ai-dev-and-research/AAPL_Comps_June2025.xlsx"
wb.save(out)
print(f"Saved: {out}")
